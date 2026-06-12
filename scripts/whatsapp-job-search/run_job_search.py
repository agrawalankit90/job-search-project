#!/usr/bin/env python3
"""
Automated Job Search Script
Triggered by GitHub Actions when WhatsApp "run job search" is received.
Searches for new PM roles, scores them, appends to recommended_jobs.xlsx, pushes to GitHub.
"""

import os
import json
import re
import subprocess
from datetime import datetime
import google.generativeai as genai
from tavily import TavilyClient
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from twilio.rest import Client as TwilioClient


# ── Config ─────────────────────────────────────────────────────────────────────
REPO_ROOT       = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
RESUME_PATH     = os.path.join(REPO_ROOT, "job-search", "my-resume.md")
STRATEGY_PATH   = os.path.join(REPO_ROOT, "job-search", "job-strategy-ankit-2026.md")
XLSX_PATH       = os.path.join(REPO_ROOT, "job-search", "recommended_jobs.xlsx")
TWILIO_FROM     = "whatsapp:+14155238886"   # Twilio sandbox number — update if using paid number


# ── Helpers ────────────────────────────────────────────────────────────────────
def read_file(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()

def get_existing_jobs(xlsx_path):
    """Return set of 'Company - Job Title' already in the tracker."""
    try:
        wb = load_workbook(xlsx_path)
        ws = wb["recommended_jobs"]
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] and row[3]:
                existing.add(f"{str(row[2]).strip()} - {str(row[3]).strip()}")
        return existing
    except Exception as e:
        print(f"Warning: could not read existing jobs: {e}")
        return set()

def search_web(tavily, queries):
    """Run multiple Tavily queries and return deduplicated results."""
    seen_urls = set()
    results = []
    for q in queries:
        try:
            res = tavily.search(query=q, search_depth="advanced", max_results=5)
            for r in res.get("results", []):
                url = r.get("url", "")
                if url not in seen_urls:
                    seen_urls.add(url)
                    results.append({
                        "title":   r.get("title", ""),
                        "url":     url,
                        "snippet": r.get("content", "")[:600],
                    })
        except Exception as e:
            print(f"Search error [{q}]: {e}")
    return results

def append_jobs_to_xlsx(new_jobs, xlsx_path):
    """Append new job rows to the recommended_jobs sheet."""
    wb = load_workbook(xlsx_path)
    ws = wb["recommended_jobs"]

    # Find current max rank
    max_rank = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and isinstance(row[0], int):
            max_rank = max(max_rank, row[0])

    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    nfont  = Font(name="Arial", size=9)
    fills  = {
        "P1": PatternFill("solid", start_color="FFF2CC"),
        "P2": PatternFill("solid", start_color="E2EFDA"),
        "P3": PatternFill("solid", start_color="EDEDED"),
        "P4": PatternFill("solid", start_color="FFE0CC"),
    }

    for job in new_jobs:
        max_rank += 1
        row_data = [
            max_rank,
            job.get("priority", "P3"),
            job.get("company", ""),
            job.get("title", ""),
            job.get("location", ""),
            job.get("job_type", "Full-time"),
            job.get("experience_level", "Senior"),
            job.get("posting_date", datetime.now().strftime("%B %Y")),
            job.get("match_pct", "65%"),
            job.get("url", ""),
            job.get("why_good_match", ""),
            job.get("gap", ""),
            job.get("rationale", ""),
        ]
        ws.append(row_data)
        last_row = ws.max_row
        prio = job.get("priority", "P3")
        for ci in range(1, 14):
            cell = ws.cell(row=last_row, column=ci)
            cell.font      = nfont
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = border
            cell.fill      = fills.get(prio, PatternFill())
        ws.row_dimensions[last_row].height = 60

    wb.save(xlsx_path)
    print(f"Appended {len(new_jobs)} new jobs to xlsx.")

def git_commit_push(message):
    """Commit updated xlsx and push to GitHub."""
    subprocess.run(["git", "config", "user.email", "ankit.sam.agrawal@gmail.com"], cwd=REPO_ROOT, check=True)
    subprocess.run(["git", "config", "user.name",  "Ankit Agrawal"],              cwd=REPO_ROOT, check=True)
    subprocess.run(["git", "add", "job-search/recommended_jobs.xlsx"],            cwd=REPO_ROOT, check=True)
    diff = subprocess.run(["git", "diff", "--staged", "--quiet"],                  cwd=REPO_ROOT)
    if diff.returncode != 0:
        subprocess.run(["git", "commit", "-m", message[:72]],                     cwd=REPO_ROOT, check=True)
        subprocess.run(["git", "push",   "origin", "main"],                       cwd=REPO_ROOT, check=True)
        print("Pushed to GitHub.")
    else:
        print("No xlsx changes to commit.")

def send_whatsapp(to_number, body):
    """Send WhatsApp message via Twilio."""
    client = TwilioClient(
        os.environ["TWILIO_ACCOUNT_SID"],
        os.environ["TWILIO_AUTH_TOKEN"]
    )
    client.messages.create(
        from_=TWILIO_FROM,
        to=f"whatsapp:{to_number}",
        body=body
    )
    print(f"WhatsApp sent to {to_number}.")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    to_number = os.environ.get("MY_WHATSAPP_NUMBER", "")

    print("Reading context files...")
    resume   = read_file(RESUME_PATH)
    strategy = read_file(STRATEGY_PATH)
    existing = get_existing_jobs(XLSX_PATH)
    existing_str = "\n".join(sorted(existing)) if existing else "None yet."

    print(f"Found {len(existing)} existing jobs. Running web search...")
    tavily = TavilyClient(api_key=os.environ["TAVILY_API_KEY"])
    queries = [
        "senior product manager jobs Bangalore India 2025 fintech AI consumer",
        "staff PM group product manager jobs India 2025 membership platform",
        "GPM product manager Bangalore 2025 consumer tech startup",
        "PM jobs India AI-native company 2025 Sarvam Krutrim",
        "senior PM jobs Bangalore 2025 Swiggy PhonePe CRED Razorpay Meesho",
        "product manager India remote 2025 GenAI agentic AI platform",
        "staff product manager India 2025 subscription billing growth",
    ]
    search_results = search_web(tavily, queries)
    search_json    = json.dumps(search_results, indent=2)
    print(f"Got {len(search_results)} search results. Calling Claude API...")

    genai.configure(api_key=os.environ["GEMINI_API_KEY"])
    client = genai.GenerativeModel("gemini-2.0-flash")
    prompt = f"""You are running an automated job search for a Staff Product Manager.

## Resume
{resume}

## Job Search Strategy
{strategy}

## Already Tracked (skip these exactly)
{existing_str}

## Live Search Results
{search_json}

## Task
1. Identify new PM job openings from the search results above.
2. Score each against the resume and strategy (only include ≥60% match).
3. Skip any job already in the tracked list (exact company + title match).
4. Assign priority: P1 ≥80% + Bangalore/India recent; P2 70–79%; P3 60–69%.

## Output
Return ONLY valid JSON — no markdown, no explanation:
{{
  "new_jobs": [
    {{
      "priority": "P1",
      "company": "Company Name",
      "title": "Job Title",
      "location": "City",
      "job_type": "Full-time",
      "experience_level": "Senior",
      "posting_date": "June 2025",
      "match_pct": "85%",
      "url": "https://...",
      "why_good_match": "Concise reason — 1-2 sentences",
      "gap": "What is missing",
      "rationale": "Priority rationale — 1 sentence"
    }}
  ],
  "summary": "Found X new jobs: Title at Company, Title at Company, ..."
}}

If no new jobs found: {{"new_jobs": [], "summary": "No new qualifying jobs found today."}}"""

    response = client.generate_content(prompt)
    result_text = response.text.strip()
    # Extract JSON — handle markdown code fences if present
    json_match = re.search(r"\{.*\}", result_text, re.DOTALL)
    if not json_match:
        summary = "Job search ran but could not parse Claude response."
        print(summary)
        if to_number:
            send_whatsapp(to_number, f"⚠️ Job search ran but hit a parsing error. Check GitHub for logs.")
        return

    result    = json.loads(json_match.group())
    new_jobs  = result.get("new_jobs", [])
    summary   = result.get("summary", f"Found {len(new_jobs)} new jobs.")

    print(f"Claude returned {len(new_jobs)} new jobs.")

    if new_jobs:
        append_jobs_to_xlsx(new_jobs, XLSX_PATH)
        git_commit_push(f"Auto job search: {summary[:60]}")

    if to_number:
        msg = (
            f"🔍 *Job Search Complete!*\n\n"
            f"{summary}\n\n"
            f"📊 View tracker: https://github.com/agrawalankit90/job-search-project/blob/main/job-search/recommended_jobs.xlsx"
        )
        send_whatsapp(to_number, msg)

    print("Job search complete.")


if __name__ == "__main__":
    main()
