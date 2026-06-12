#!/usr/bin/env python3
"""
Automated HR Contact Search Script
Triggered by GitHub Actions when WhatsApp "run hr search" is received.
Finds hiring managers / recruiters for all P1+P2 unapplied jobs,
updates the HR_repository sheet in recommended_jobs.xlsx, pushes to GitHub.
No LLM required — extracts contact info directly from Tavily search results.
"""

import os
import re
import subprocess
from datetime import datetime
from tavily import TavilyClient
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from twilio.rest import Client as TwilioClient


# ── Config ─────────────────────────────────────────────────────────────────────
REPO_ROOT   = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX_PATH   = os.path.join(REPO_ROOT, "job-search", "recommended_jobs.xlsx")
TWILIO_FROM = "whatsapp:+14155238886"

# Patterns to extract LinkedIn profile URLs
LINKEDIN_RE = re.compile(r"https?://(?:www\.)?linkedin\.com/in/[A-Za-z0-9\-_%]+/?")

# Common name patterns in LinkedIn/recruiter snippets
NAME_RE = re.compile(
    r"\b([A-Z][a-z]+(?:\s[A-Z][a-z]+)+)"  # Two+ capitalised words
)


# ── Helpers ────────────────────────────────────────────────────────────────────
def get_unapplied_p1_p2(xlsx_path):
    """Return list of dicts for P1/P2 jobs with no 'applied' status."""
    wb  = load_workbook(xlsx_path)
    ws  = wb["recommended_jobs"]
    jobs = []
    headers = [cell.value for cell in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h}

    for row in ws.iter_rows(min_row=2, values_only=True):
        priority   = str(row[idx.get("Priority", 1)] or "").strip()
        status_idx = idx.get("Status", 13)
        status     = str(row[status_idx] if status_idx < len(row) else "").strip().lower()
        if priority in ("P1", "P2") and status != "applied":
            jobs.append({
                "company":  str(row[idx.get("Company Name", 2)] or "").strip(),
                "title":    str(row[idx.get("Job Title",    3)] or "").strip(),
                "location": str(row[idx.get("Location",     4)] or "").strip(),
                "priority": priority,
            })
    return jobs

def get_existing_hr_contacts(xlsx_path):
    """Return set of 'Title|Company' already in HR_repository."""
    try:
        wb = load_workbook(xlsx_path)
        if "HR_repository" not in wb.sheetnames:
            return set()
        ws = wb["HR_repository"]
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                existing.add(f"{str(row[0]).strip()}|{str(row[1]).strip()}")
        return existing
    except Exception:
        return set()

def search_contacts(tavily, company, title):
    """Run Tavily searches and return deduplicated results."""
    queries = [
        f'"{company}" recruiter "talent acquisition" product manager LinkedIn',
        f'"{company}" "hiring manager" "{title}" site:linkedin.com',
        f'"{company}" HR recruiter email contact product manager',
    ]
    seen_urls = set()
    results   = []
    for q in queries:
        try:
            res = tavily.search(query=q, search_depth="basic", max_results=3)
            for r in res.get("results", []):
                url = r.get("url", "")
                if url not in seen_urls:
                    seen_urls.add(url)
                    results.append({
                        "title":   r.get("title", ""),
                        "url":     url,
                        "snippet": r.get("content", "")[:600],
                    })
        except Exception as e:
            print(f"  Search error [{q[:50]}]: {e}")
    return results

def extract_contact(results, company):
    """
    Extract contact info directly from Tavily snippets — no LLM needed.
    Returns a dict with contact_name, contact_role, email, linkedin_url,
    source, confidence, notes.
    """
    contact_name  = ""
    contact_role  = ""
    email         = ""
    linkedin_url  = ""
    source        = "Tavily Search"
    confidence    = "Low"
    notes         = "LinkedIn InMail recommended as primary outreach channel."

    all_text = " ".join(
        f"{r['title']} {r['snippet']}" for r in results
    )
    all_urls = [r["url"] for r in results]

    # 1. Extract LinkedIn URL
    li_matches = LINKEDIN_RE.findall(" ".join(all_urls + [all_text]))
    if li_matches:
        linkedin_url = li_matches[0]
        source       = "LinkedIn"
        confidence   = "Medium"

    # 2. Extract email
    email_match = re.search(
        r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", all_text
    )
    if email_match:
        email      = email_match.group()
        confidence = "High" if linkedin_url else "Medium"
        source     = "LinkedIn + email" if linkedin_url else "Web search"

    # 3. Extract name — look for recruiter/TA/HR context
    recruiter_keywords = ["recruiter", "talent acquisition", "hiring manager",
                          "hr manager", "people operations", "head of talent"]
    for keyword in recruiter_keywords:
        pattern = re.compile(
            rf"([A-Z][a-z]+(?:\s[A-Z][a-z]+)+)\s*[,\-–]?\s*{keyword}",
            re.IGNORECASE
        )
        match = pattern.search(all_text)
        if match:
            contact_name = match.group(1).strip()
            contact_role = keyword.title()
            break

    # 4. Fallback: first capitalised name near company name
    if not contact_name:
        company_pattern = re.compile(
            rf"{re.escape(company)}.{{0,80}}?([A-Z][a-z]+ [A-Z][a-z]+)",
            re.IGNORECASE | re.DOTALL
        )
        match = company_pattern.search(all_text)
        if match:
            contact_name = match.group(1).strip()
            contact_role = "Recruiter / TA (unverified)"
            confidence   = "Low"

    # 5. Build notes
    if confidence == "Low":
        notes = "No verified contact found — apply via company careers page or LinkedIn InMail."
    elif confidence == "Medium":
        notes = "LinkedIn profile found. InMail recommended; email not verified."
    else:
        notes = "Email and LinkedIn found. Verify before outreach."

    return {
        "contact_name": contact_name or "Not found",
        "contact_role": contact_role,
        "email":        email,
        "linkedin_url": linkedin_url,
        "source":       source,
        "confidence":   confidence,
        "notes":        notes,
    }

def upsert_hr_row(wb, row_data):
    """Insert or update a row in HR_repository sheet."""
    if "HR_repository" not in wb.sheetnames:
        _create_hr_sheet(wb)
    ws = wb["HR_repository"]

    key = f"{row_data[0]}|{row_data[1]}"
    for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
        existing_key = f"{str(row[0].value or '').strip()}|{str(row[1].value or '').strip()}"
        if existing_key == key:
            for ci, val in enumerate(row_data, 1):
                ws.cell(row=ri, column=ci, value=val)
            _style_hr_row(ws, ri, row_data[9])
            return

    ws.append(row_data)
    _style_hr_row(ws, ws.max_row, row_data[9])

def _create_hr_sheet(wb):
    ws         = wb.create_sheet("HR_repository")
    headers    = ["Job Title","Company","Priority","Location","Contact Name",
                  "Contact Role","Email Address","LinkedIn URL","Source","Confidence Level","Notes"]
    col_widths = [35,16,8,22,22,32,32,50,28,16,45]
    hfill  = PatternFill("solid", start_color="1F4E79")
    hfont  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    from openpyxl.utils import get_column_letter
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hfont
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

def _style_hr_row(ws, row_num, confidence):
    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    nfont  = Font(name="Arial", size=9)
    conf_fills = {
        "High":   PatternFill("solid", start_color="C6EFCE"),
        "Medium": PatternFill("solid", start_color="FFEB9C"),
        "Low":    PatternFill("solid", start_color="FFC7CE"),
    }
    for ci in range(1, 12):
        cell = ws.cell(row=row_num, column=ci)
        cell.font      = nfont
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border    = border
        if ci == 10:
            cell.fill = conf_fills.get(str(confidence), PatternFill())
    ws.row_dimensions[row_num].height = 60

def git_commit_push(message):
    subprocess.run(["git", "config", "user.email", "ankit.sam.agrawal@gmail.com"], cwd=REPO_ROOT, check=True)
    subprocess.run(["git", "config", "user.name",  "Ankit Agrawal"],              cwd=REPO_ROOT, check=True)
    subprocess.run(["git", "add", "job-search/recommended_jobs.xlsx"],            cwd=REPO_ROOT, check=True)
    diff = subprocess.run(["git", "diff", "--staged", "--quiet"], cwd=REPO_ROOT)
    if diff.returncode != 0:
        subprocess.run(["git", "commit", "-m", message[:72]], cwd=REPO_ROOT, check=True)
        subprocess.run(["git", "push",   "origin", "main"],   cwd=REPO_ROOT, check=True)
        print("Pushed to GitHub.")
    else:
        print("No changes to commit.")

def send_whatsapp(to_number, body):
    client = TwilioClient(os.environ["TWILIO_ACCOUNT_SID"], os.environ["TWILIO_AUTH_TOKEN"])
    client.messages.create(from_=TWILIO_FROM, to=f"whatsapp:{to_number}", body=body)
    print(f"WhatsApp sent to {to_number}.")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    to_number = os.environ.get("MY_WHATSAPP_NUMBER", "")

    print("Reading P1/P2 unapplied jobs from xlsx...")
    jobs     = get_unapplied_p1_p2(XLSX_PATH)
    existing = get_existing_hr_contacts(XLSX_PATH)
    print(f"Found {len(jobs)} target roles.")

    tavily = TavilyClient(api_key=os.environ["TAVILY_API_KEY"])
    wb     = load_workbook(XLSX_PATH)

    updated = 0
    for job in jobs:
        key = f"{job['title']}|{job['company']}"
        print(f"Searching contacts for: {job['company']} — {job['title']}")
        search_results = search_contacts(tavily, job["company"], job["title"])
        contact        = extract_contact(search_results, job["company"])

        row_data = [
            job["title"],
            job["company"],
            job["priority"],
            job["location"],
            contact["contact_name"],
            contact["contact_role"],
            contact["email"],
            contact["linkedin_url"],
            contact["source"],
            contact["confidence"],
            contact["notes"],
        ]
        upsert_hr_row(wb, row_data)
        updated += 1
        print(f"  ✓ {contact['contact_name']} — {contact['confidence']} confidence")

    wb.save(XLSX_PATH)
    git_commit_push(f"Auto HR search: updated contacts for {updated} roles")

    summary = f"Found/updated contacts for {updated} of {len(jobs)} roles."
    if to_number:
        msg = (
            f"📋 *HR Contact Search Complete!*\n\n"
            f"{summary}\n\n"
            f"📊 View HR_repository: https://github.com/agrawalankit90/job-search-project/blob/main/job-search/recommended_jobs.xlsx"
        )
        send_whatsapp(to_number, msg)

    print(f"HR search complete. {summary}")


if __name__ == "__main__":
    main()
